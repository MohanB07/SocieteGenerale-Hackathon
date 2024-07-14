import requests
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

API_KEY = 'TY8T0XID9SPIUF4P'
SYMBOL = 'SCGLY'

def fetch_stock_data(symbol, interval='daily', outputsize='compact'):
    """
    Fetches historical stock data from Alpha Vantage API.
    Returns a DataFrame with daily stock data.
    """
    url = 'https://www.alphavantage.co/query'
    params = {
        'function': 'TIME_SERIES_DAILY',
        'symbol': symbol,
        'apikey': API_KEY,
        'outputsize': outputsize
    }

    response = requests.get(url, params=params)
    data = response.json()

    if 'Time Series (Daily)' in data:
        df = pd.DataFrame(data['Time Series (Daily)']).T
        df.index = pd.to_datetime(df.index)
        df.sort_index(inplace=True)
        df.columns = ['Open', 'High', 'Low', 'Close', 'Volume']
        df = df.astype(float)
        return df
    else:
        print(f"Error fetching data: {data['Error Message']}")
        return None

def update_excel_with_stock_data(df, sheet_name='Data'):
    """
    Updates the Excel sheet with stock data.
    """
    try:
        excel_file = 'FinancialData.xlsx'
        wb = load_workbook(excel_file)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(2, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        wb.save(excel_file)
        wb.close()
        print(f"Updated {sheet_name} sheet with new data.")

    except Exception as e:
        print(f"Error updating Excel sheet: {str(e)}")

def preload_past_two_months_data():
    """
    Preloads data for the past two months for initial setup/testing.
    """
    today = datetime.today()
    start_date = today - timedelta(days=60)  # 60 days ago
    end_date = today

    df = fetch_stock_data(SYMBOL, interval='daily', outputsize='full')
    if df is not None:
        df = df.loc[start_date:end_date]
        update_excel_with_stock_data(df, sheet_name='Data')

def main():
    """
    Main function to fetch and update daily data.
    """
    preload_past_two_months_data()

    # Schedule daily update (replace with your scheduling method)
    while True:
        now = datetime.now()
        if now.hour == 0 and now.minute == 0:  # Update daily at midnight
            df_today = fetch_stock_data(SYMBOL, interval='daily', outputsize='compact')
            if df_today is not None:
                update_excel_with_stock_data(df_today, sheet_name='Data')
        # Sleep for 1 hour before checking again
        time.sleep(3600)

if __name__ == '__main__':
    main()
